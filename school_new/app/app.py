from flask import Flask,render_template,request,redirect,url_for,flash,session,send_file,abort
import sqlite3, shutil, tempfile, os, csv
from pathlib import Path
from datetime import date, datetime
from functools import wraps
from urllib.parse import quote
import json, urllib.request, urllib.error, secrets
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook, Workbook
from io import BytesIO, StringIO

BASE=Path(__file__).resolve().parent; DB=BASE/'school.db'; BACKUPS=BASE/'backups'; BACKUPS.mkdir(exist_ok=True)
INV_UPLOADS=BASE/'invitation_uploads'; INV_UPLOADS.mkdir(exist_ok=True)
ALLOWED_INV_FILES={'png','jpg','jpeg','webp','pdf','doc','docx'}
MAX_INV_FILE_SIZE=10*1024*1024
app=Flask(__name__); app.secret_key=os.environ.get('SCHOOL_SECRET_KEY','school-management-change-this-secret')
ROLES={'admin':'مدیر سیستم','operator':'کاربر مدرسه','viewer':'فقط مشاهده'}

# ---------- Persian/Jalali date ----------
def conn():
 c=sqlite3.connect(DB); c.row_factory=sqlite3.Row; c.execute('PRAGMA foreign_keys=ON'); return c

def gy_to_jalali(gy,gm,gd):
 gdm=[31,28,31,30,31,30,31,31,30,31,30,31]
 gy2=gy-1600; gdn=365*gy2+(gy2+3)//4-(gy2+99)//100+(gy2+399)//400
 for i in range(gm-1): gdn+=gdm[i]
 if gm>2 and (gy%4==0 and gy%100!=0 or gy%400==0): gdn+=1
 gdn+=gd-1; jdn=gdn-79; jnp=jdn//12053; jdn%=12053
 jy=979+33*jnp+4*(jdn//1461); jdn%=1461
 if jdn>=366: jy+=(jdn-1)//365; jdn=(jdn-1)%365
 if jdn<186: jm=1+jdn//31; jd=1+jdn%31
 else: jm=7+(jdn-186)//30; jd=1+(jdn-186)%30
 return jy,jm,jd

def jalali_to_gy(jy,jm,jd):
 jy-=979; jm-=1; jd-=1; jdn=365*jy+(jy//33)*8+((jy%33)+3)//4
 for i in range(jm): jdn+=31 if i<6 else 30
 jdn+=jd; gdn=jdn+79; gy=1600+400*(gdn//146097); gdn%=146097; leap=True
 if gdn>=36525:
  gdn-=1; gy+=100*(gdn//36524); gdn%=36524
  if gdn>=365: gdn+=1
  else: leap=False
 gy+=4*(gdn//1461); gdn%=1461
 if gdn>=366: leap=False; gy+=(gdn-1)//365; gdn=(gdn-1)%365
 md=[31,29 if leap else 28,31,30,31,30,31,31,30,31,30,31]; gm=1
 while gdn>=md[gm-1]: gdn-=md[gm-1]; gm+=1
 return gy,gm,gdn+1

def today_j():
 d=date.today(); y,m,dd=gy_to_jalali(d.year,d.month,d.day); return f'{y:04d}/{m:02d}/{dd:02d}'
def norm(s): return (s or '').translate(str.maketrans('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩','01234567890123456789'))
def j_to_iso(s):
 s=norm(str(s or '')).strip().replace('\u200c','').replace('\u200f','').replace('\u200e','')
 s=s.replace('-', '/').replace('.', '/').replace('\\','/')
 if s.endswith('.0'): s=s[:-2]
 if s.isdigit() and len(s)==8: s=f'{s[:4]}/{s[4:6]}/{s[6:8]}'
 parts=[x for x in s.split('/') if x!='']
 if len(parts)!=3: raise ValueError('تاریخ باید به شکل 1395/05/26 یا 13950526 باشد.')
 y,m,d=map(int,parts)
 if not (1300 <= y <= 1500 and 1 <= m <= 12 and 1 <= d <= 31): raise ValueError('سال/ماه/روز نامعتبر است.')
 gy,gm,gd=jalali_to_gy(y,m,d); return f'{gy:04d}-{gm:02d}-{gd:02d}'

def excel_text(v):
 if v is None: return ''
 if isinstance(v,float) and v.is_integer(): return str(int(v))
 return norm(str(v)).strip()

def normalize_phone(v):
 x=excel_text(v).replace(' ','').replace('-','').replace('(','').replace(')','')
 if x.endswith('.0'): x=x[:-2]
 # Excel often strips the leading zero from Iranian mobile numbers.
 if x.isdigit() and len(x)==10 and x.startswith('9'): x='0'+x
 return x

def iso_to_j(s):
 try:
  y,m,d=map(int,str(s)[:10].split('-')); y,m,d=gy_to_jalali(y,m,d); return f'{y:04d}/{m:02d}/{d:02d}'
 except: return s or ''
@app.template_filter('jdate')
def jdate(s): return iso_to_j(s)
@app.context_processor
def ctx(): return {'current_user':session.get('username'),'current_role':ROLES.get(session.get('role'),''),'today_jalali':today_j(),'roles':ROLES}

# ---------- Database ----------
def init_db():
 c=conn(); c.executescript('''
 CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY,username TEXT UNIQUE NOT NULL,password_hash TEXT NOT NULL,role TEXT NOT NULL,active INTEGER DEFAULT 1,must_change_password INTEGER DEFAULT 0,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
 CREATE TABLE IF NOT EXISTS students(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT NOT NULL,code TEXT UNIQUE NOT NULL,national_id TEXT,class_name TEXT,grade TEXT,phone TEXT,guardian_phone TEXT,father_name TEXT,father_phone TEXT,mother_phone TEXT,whatsapp TEXT,bale TEXT,rubika TEXT,eitaa TEXT,birth_date TEXT,birthday_approved_year TEXT,notes TEXT,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
 CREATE TABLE IF NOT EXISTS attendance(id INTEGER PRIMARY KEY,student_id INTEGER,att_date TEXT,status TEXT,UNIQUE(student_id,att_date),FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE);
 CREATE TABLE IF NOT EXISTS discipline(id INTEGER PRIMARY KEY,student_id INTEGER,event_date TEXT,event_type TEXT,action TEXT,description TEXT,FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE);
 CREATE TABLE IF NOT EXISTS exams(id INTEGER PRIMARY KEY AUTOINCREMENT,title TEXT NOT NULL,grade TEXT,class_name TEXT,start_date TEXT,end_date TEXT,duration_minutes INTEGER DEFAULT 30,phone_required INTEGER DEFAULT 0,phone_target TEXT DEFAULT 'student',random_questions INTEGER DEFAULT 0,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP,public_token TEXT UNIQUE NOT NULL);
 CREATE TABLE IF NOT EXISTS questions(id INTEGER PRIMARY KEY AUTOINCREMENT,exam_id INTEGER,question_text TEXT NOT NULL,qtype TEXT DEFAULT 'mcq',option_a TEXT,option_b TEXT,option_c TEXT,option_d TEXT,correct_answer TEXT,points REAL DEFAULT 1,sort_order INTEGER DEFAULT 0,FOREIGN KEY(exam_id) REFERENCES exams(id) ON DELETE CASCADE);
 CREATE TABLE IF NOT EXISTS attempts(id INTEGER PRIMARY KEY AUTOINCREMENT,exam_id INTEGER,student_id INTEGER,phone TEXT,started_at TEXT,submitted_at TEXT,score REAL DEFAULT 0,status TEXT DEFAULT 'started',UNIQUE(exam_id,student_id),FOREIGN KEY(exam_id) REFERENCES exams(id) ON DELETE CASCADE,FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE);
 CREATE TABLE IF NOT EXISTS answers(id INTEGER PRIMARY KEY AUTOINCREMENT,attempt_id INTEGER,question_id INTEGER,answer TEXT,points REAL DEFAULT 0,FOREIGN KEY(attempt_id) REFERENCES attempts(id) ON DELETE CASCADE,FOREIGN KEY(question_id) REFERENCES questions(id) ON DELETE CASCADE);
 CREATE TABLE IF NOT EXISTS messaging_config(id INTEGER PRIMARY KEY CHECK(id=1),whatsapp_enabled INTEGER DEFAULT 0,whatsapp_token TEXT DEFAULT '',whatsapp_phone_number_id TEXT DEFAULT '',whatsapp_api_version TEXT DEFAULT 'v23.0',whatsapp_template_name TEXT DEFAULT '',whatsapp_language TEXT DEFAULT 'fa');
 CREATE TABLE IF NOT EXISTS message_logs(id INTEGER PRIMARY KEY AUTOINCREMENT,student_id INTEGER,recipient TEXT,platform TEXT,message TEXT,status TEXT,detail TEXT,created_at TEXT DEFAULT CURRENT_TIMESTAMP,FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE SET NULL);
 CREATE TABLE IF NOT EXISTS invitation_files(id INTEGER PRIMARY KEY AUTOINCREMENT,token TEXT UNIQUE NOT NULL,original_name TEXT NOT NULL,stored_name TEXT NOT NULL,mime_type TEXT,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
 ''')
 cols={r['name'] for r in c.execute('PRAGMA table_info(students)')}
 for col in ['national_id','father_name','father_phone','mother_phone','father_whatsapp','father_bale','father_rubika','father_eitaa','mother_whatsapp','mother_bale','mother_rubika','mother_eitaa','birth_date','birthday_approved_year']:
  if col not in cols: c.execute(f'ALTER TABLE students ADD COLUMN {col} TEXT')
 if c.execute('SELECT COUNT(*) n FROM users').fetchone()['n']==0: c.execute('INSERT INTO users(username,password_hash,role,must_change_password) VALUES(?,?,?,1)',('admin',generate_password_hash('admin123'),'admin'))
 c.commit(); c.close()
init_db()

# ---------- Auth ----------
def login_required(f):
 @wraps(f)
 def w(*a,**k):
  if 'user_id' not in session:return redirect(url_for('login',next=request.path))
  return f(*a,**k)
 return w
def perm(*roles):
 def deco(f):
  @wraps(f)
  def w(*a,**k):
   if 'user_id' not in session:return redirect(url_for('login'))
   if session.get('role') not in roles: flash('دسترسی شما به این بخش مجاز نیست.','danger'); return redirect(url_for('dashboard'))
   return f(*a,**k)
  return w
 return deco
@app.route('/login',methods=['GET','POST'])
def login():
 if request.method=='POST':
  c=conn(); u=c.execute('SELECT * FROM users WHERE username=? AND active=1',(request.form['username'].strip(),)).fetchone(); c.close()
  if u and check_password_hash(u['password_hash'],request.form['password']):
   session.clear(); session.update(user_id=u['id'],username=u['username'],role=u['role'],must_change_password=bool(u['must_change_password']))
   return redirect(url_for('change_password') if u['must_change_password'] else url_for('dashboard'))
  flash('نام کاربری یا رمز عبور نادرست است.','danger')
 return render_template('login.html')
@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))
@app.route('/change-password',methods=['GET','POST'])
@login_required
def change_password():
 if request.method=='POST':
  c=conn(); u=c.execute('SELECT * FROM users WHERE id=?',(session['user_id'],)).fetchone(); new=request.form['new_password']
  if not check_password_hash(u['password_hash'],request.form['old_password']): flash('رمز فعلی نادرست است.','danger')
  elif len(new)<6: flash('رمز جدید حداقل ۶ کاراکتر باشد.','danger')
  elif new!=request.form['repeat_password']: flash('تکرار رمز یکسان نیست.','danger')
  else: c.execute('UPDATE users SET password_hash=?,must_change_password=0 WHERE id=?',(generate_password_hash(new),u['id'])); c.commit(); session['must_change_password']=False; flash('رمز با موفقیت تغییر کرد.','success'); c.close(); return redirect(url_for('dashboard'))
  c.close()
 return render_template('change_password.html')
@app.before_request
def force_change():
 if request.endpoint not in ('login','static','change_password') and session.get('must_change_password'): return redirect(url_for('change_password'))

# ---------- Dashboard / birthdays ----------
def jalali_birthday_candidates(max_days=3):
 # Returns students whose next birthday is today or within the next max_days.
 # The exact 3-day reminder is highlighted separately.
 now_j=today_j()
 jy=int(now_j[:4])
 today_g=date.today()
 c=conn(); rows=c.execute("SELECT * FROM students WHERE birth_date IS NOT NULL AND birth_date<>'' ORDER BY name").fetchall(); c.close()
 out=[]
 for s in rows:
  try:
   by,bm,bd=map(int,iso_to_j(s['birth_date']).split('/'))
   target=date.fromisoformat(j_to_iso(f'{jy:04d}/{bm:02d}/{bd:02d}'))
   if target < today_g:
    target=date.fromisoformat(j_to_iso(f'{jy+1:04d}/{bm:02d}/{bd:02d}'))
   delta=(target-today_g).days
   if 0 <= delta <= max_days:
    item=dict(s)
    item['days_until']=delta
    item['reminder_text']='امروز تولد است' if delta==0 else f'{delta} روز مانده تا تولد'
    out.append(item)
  except Exception:
   continue
 return out

def birthday_rows():
 return [x for x in jalali_birthday_candidates(0)]

def birthday_reminders():
 return jalali_birthday_candidates(3)

@app.route('/')
@login_required
def dashboard():
 c=conn(); total=c.execute('SELECT COUNT(*) n FROM students').fetchone()['n']; d=date.today().isoformat(); present=c.execute("SELECT COUNT(*) n FROM attendance WHERE att_date=? AND status='حاضر'",(d,)).fetchone()['n']; absent=c.execute("SELECT COUNT(*) n FROM attendance WHERE att_date=? AND status='غایب'",(d,)).fetchone()['n']; late=c.execute("SELECT COUNT(*) n FROM attendance WHERE att_date=? AND status='تأخیر'",(d,)).fetchone()['n']; dis=c.execute('SELECT COUNT(*) n FROM discipline').fetchone()['n']; c.close(); return render_template('dashboard.html',total=total,present=present,absent=absent,late=late,discipline=dis,birthdays=birthday_rows(),birthday_reminders=birthday_reminders())


# ---------- Messaging helpers ----------
def get_messaging_config():
 c=conn(); row=c.execute('SELECT * FROM messaging_config WHERE id=1').fetchone(); c.close(); return row

def save_messaging_config(d):
 c=conn(); c.execute("INSERT INTO messaging_config(id,whatsapp_enabled,whatsapp_token,whatsapp_phone_number_id,whatsapp_api_version,whatsapp_template_name,whatsapp_language) VALUES(1,?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET whatsapp_enabled=excluded.whatsapp_enabled,whatsapp_token=excluded.whatsapp_token,whatsapp_phone_number_id=excluded.whatsapp_phone_number_id,whatsapp_api_version=excluded.whatsapp_api_version,whatsapp_template_name=excluded.whatsapp_template_name,whatsapp_language=excluded.whatsapp_language",(int(d.get('whatsapp_enabled','0')),d.get('whatsapp_token','').strip(),d.get('whatsapp_phone_number_id','').strip(),d.get('whatsapp_api_version','v23.0').strip() or 'v23.0',d.get('whatsapp_template_name','').strip(),d.get('whatsapp_language','fa').strip() or 'fa')); c.commit(); c.close()

def send_whatsapp(to_number, text):
 cfg=get_messaging_config()
 if not cfg or not cfg['whatsapp_enabled'] or not cfg['whatsapp_token'] or not cfg['whatsapp_phone_number_id'] or not cfg['whatsapp_template_name']:
  return False,'API واتساپ تنظیم نشده است.'
 p=phone(to_number); p='98'+p[1:] if p.startswith('0') else p
 payload={'messaging_product':'whatsapp','to':p,'type':'template','template':{'name':cfg['whatsapp_template_name'],'language':{'code':cfg['whatsapp_language']},'components':[{'type':'body','parameters':[{'type':'text','text':text}]}]}}
 url=f"https://graph.facebook.com/{cfg['whatsapp_api_version']}/{cfg['whatsapp_phone_number_id']}/messages"
 try:
  req=urllib.request.Request(url,data=json.dumps(payload).encode(),headers={'Authorization':'Bearer '+cfg['whatsapp_token'],'Content-Type':'application/json'},method='POST')
  with urllib.request.urlopen(req,timeout=20) as r: data=json.loads(r.read().decode())
  return True,data.get('messages',[{}])[0].get('id','sent')
 except Exception as e: return False,str(e)


def send_whatsapp_image(to_number, file_path, caption=''):
 cfg=get_messaging_config()
 if not cfg or not cfg['whatsapp_enabled'] or not cfg['whatsapp_token'] or not cfg['whatsapp_phone_number_id']:
  return False,'API رسمی واتساپ تنظیم نشده است.'
 p=phone(to_number); p='98'+p[1:] if p.startswith('0') else p
 try:
  path=Path(file_path)
  mime={'.jpg':'image/jpeg','.jpeg':'image/jpeg','.png':'image/png','.webp':'image/webp'}.get(path.suffix.lower())
  if not mime: return False,'فقط تصویر JPG، PNG یا WEBP برای ارسال تصویری پشتیبانی می‌شود.'
  boundary='----SchoolManagementBoundary'+secrets.token_hex(12)
  data=path.read_bytes()
  body=(f'--{boundary}\r\nContent-Disposition: form-data; name="messaging_product"\r\n\r\nwhatsapp\r\n').encode()
  body+=(f'--{boundary}\r\nContent-Disposition: form-data; name="file"; filename="{path.name}"\r\nContent-Type: {mime}\r\n\r\n').encode()+data+b'\r\n'+(f'--{boundary}--\r\n').encode()
  base=f"https://graph.facebook.com/{cfg['whatsapp_api_version']}"
  req=urllib.request.Request(f"{base}/{cfg['whatsapp_phone_number_id']}/media",data=body,headers={'Authorization':'Bearer '+cfg['whatsapp_token'],'Content-Type':f'multipart/form-data; boundary={boundary}'},method='POST')
  with urllib.request.urlopen(req,timeout=30) as r: media=json.loads(r.read().decode())
  media_id=media.get('id')
  if not media_id: return False,'شناسه تصویر از واتساپ دریافت نشد.'
  payload={'messaging_product':'whatsapp','to':p,'type':'image','image':{'id':media_id}}
  if caption: payload['image']['caption']=caption[:1024]
  req=urllib.request.Request(f"{base}/{cfg['whatsapp_phone_number_id']}/messages",data=json.dumps(payload).encode(),headers={'Authorization':'Bearer '+cfg['whatsapp_token'],'Content-Type':'application/json'},method='POST')
  with urllib.request.urlopen(req,timeout=30) as r: out=json.loads(r.read().decode())
  return True,out.get('messages',[{}])[0].get('id','sent')
 except Exception as e: return False,str(e)

def send_or_prepare(student, recipient, platform, text):
 num=student['father_phone'] if recipient=='پدر' else student['mother_phone'] if recipient=='مادر' else student['phone']
 if platform=='whatsapp' and num:
  ok,detail=send_whatsapp(num,text)
  c=conn(); c.execute('INSERT INTO message_logs(student_id,recipient,platform,message,status,detail) VALUES(?,?,?,?,?,?)',(student['id'],recipient,platform,text,'sent' if ok else 'failed',detail)); c.commit(); c.close(); return ok,detail
 return False,'API رسمی این پیام‌رسان در این نسخه فعال نشده است.'

# ---------- Students ----------
@app.route('/students')
@login_required
def students():
 q=request.args.get('q','').strip(); grade=request.args.get('grade','').strip(); class_name=request.args.get('class_name','').strip()
 # Persian-safe search: SQLite LIKE is not reliable for Arabic/Persian character variants.
 def search_norm(v):
  # Robust Persian/Arabic normalization for names and numbers.
  import unicodedata
  x=norm(str(v or ''))
  x=unicodedata.normalize('NFKC', x)
  x=(x.replace('ي','ی').replace('ى','ی').replace('ئ','ی')
       .replace('ك','ک').replace('ة','ه').replace('ۀ','ه'))
  x=x.replace('ـ','')
  # Remove zero-width chars and Arabic vowel marks/diacritics.
  x=x.replace('\u200c','').replace('\u200f','').replace('\u200e','').replace('\ufeff','')
  x=''.join(ch for ch in x if unicodedata.category(ch) not in ('Mn','Cf'))
  x=' '.join(x.split())
  return x.casefold()
 nq=search_norm(q); ng=search_norm(grade); nc=search_norm(class_name)
 c=conn(); all_rows=c.execute('SELECT * FROM students ORDER BY grade,class_name,name').fetchall()
 grades=c.execute("SELECT DISTINCT grade FROM students WHERE grade<>'' ORDER BY grade").fetchall()
 classes=c.execute("SELECT DISTINCT class_name FROM students WHERE class_name<>'' ORDER BY class_name").fetchall(); c.close()
 rows=[]
 for s in all_rows:
  if ng and search_norm(s['grade']) != ng: continue
  if nc and search_norm(s['class_name']) != nc: continue
  if nq:
   fields=[s['name'],s['code'],s['national_id'],s['class_name'],s['grade'],s['phone'],s['guardian_phone'],s['father_name'],s['father_phone'],s['mother_phone'],s['whatsapp'],s['bale'],s['rubika'],s['eitaa'],s['father_whatsapp'],s['father_bale'],s['father_rubika'],s['father_eitaa'],s['mother_whatsapp'],s['mother_bale'],s['mother_rubika'],s['mother_eitaa']]
   # Match both with and without spaces, so 'آرتا تلخ ابی' and 'آرتا تلخابی' work.
   nq_compact=nq.replace(' ','')
   if not any(nq in search_norm(v) or nq_compact in search_norm(v).replace(' ','') for v in fields): continue
  rows.append(s)
 return render_template('students.html',students=rows,q=q,grade=grade,class_name=class_name,grades=grades,classes=classes)
@app.route('/students/add',methods=['GET','POST'])
@perm('admin','operator')
def add_student(): return student_edit(None)
@app.route('/students/<int:sid>/edit',methods=['GET','POST'])
@perm('admin','operator')
def edit_student(sid): return student_edit(sid)
def student_edit(sid):
 c=conn(); s=c.execute('SELECT * FROM students WHERE id=?',(sid,)).fetchone() if sid else None
 if request.method=='POST':
  d=request.form; birth=''
  if d.get('birth_date','').strip():
   try: birth=j_to_iso(d['birth_date'])
   except: flash('تاریخ تولد شمسی نامعتبر است.','danger'); c.close(); return render_template('student_form.html',student=s)
  vals=(d['name'],d['code'],d.get('national_id',''),d.get('grade',''),d.get('class_name',''),d.get('phone',''),d.get('guardian_phone',''),d.get('father_name',''),d.get('father_phone',''),d.get('mother_phone',''),d.get('father_whatsapp',''),d.get('father_bale',''),d.get('father_rubika',''),d.get('father_eitaa',''),d.get('mother_whatsapp',''),d.get('mother_bale',''),d.get('mother_rubika',''),d.get('mother_eitaa',''),d.get('whatsapp',''),d.get('bale',''),d.get('rubika',''),d.get('eitaa',''),birth,d.get('notes',''))
  try:
   if sid: c.execute('UPDATE students SET name=?,code=?,national_id=?,grade=?,class_name=?,phone=?,guardian_phone=?,father_name=?,father_phone=?,mother_phone=?,father_whatsapp=?,father_bale=?,father_rubika=?,father_eitaa=?,mother_whatsapp=?,mother_bale=?,mother_rubika=?,mother_eitaa=?,whatsapp=?,bale=?,rubika=?,eitaa=?,birth_date=?,notes=? WHERE id=?',vals+(sid,))
   else: c.execute('INSERT INTO students(name,code,national_id,grade,class_name,phone,guardian_phone,father_name,father_phone,mother_phone,father_whatsapp,father_bale,father_rubika,father_eitaa,mother_whatsapp,mother_bale,mother_rubika,mother_eitaa,whatsapp,bale,rubika,eitaa,birth_date,notes) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',vals)
   c.commit(); c.close(); flash('اطلاعات ذخیره شد.','success'); return redirect(url_for('students'))
  except sqlite3.IntegrityError: flash('کد دانش‌آموزی تکراری است.','danger')
 c.close(); return render_template('student_form.html',student=s)
@app.post('/students/<int:sid>/delete')
@perm('admin')
def delete_student(sid): c=conn(); c.execute('DELETE FROM students WHERE id=?',(sid,)); c.commit(); c.close(); flash('دانش‌آموز حذف شد.','success'); return redirect(url_for('students'))
@app.post('/students/bulk-delete')
@perm('admin')
def bulk_delete_students():
 ids=[int(x) for x in request.form.getlist('student_ids') if str(x).isdigit()]
 delete_all=request.form.get('delete_all')=='1'
 c=conn()
 if delete_all:
  q=request.form.get('q','').strip(); grade=request.form.get('grade','').strip(); class_name=request.form.get('class_name','').strip()
  rows=c.execute('SELECT id FROM students WHERE (name LIKE ? OR code LIKE ? OR class_name LIKE ? OR grade LIKE ? OR phone LIKE ? OR guardian_phone LIKE ? OR father_phone LIKE ? OR mother_phone LIKE ?) AND (?='' OR grade=?) AND (?='' OR class_name=?)',(f'%{q}%',)*8+(grade,grade,class_name,class_name)).fetchall()
  ids=[r['id'] for r in rows]
 if not ids:
  c.close(); flash('هیچ دانش‌آموزی برای حذف انتخاب نشده است.','warning'); return redirect(url_for('students'))
 c.executemany('DELETE FROM students WHERE id=?',[(i,) for i in ids]); c.commit(); c.close(); flash(f'{len(ids)} دانش‌آموز حذف شد.','success'); return redirect(url_for('students'))

# ---------- Excel ----------
@app.route('/students/import',methods=['GET','POST'])
@perm('admin','operator')
def import_students():
 result=None
 if request.method=='POST':
  f=request.files.get('excel')
  try:
   wb=load_workbook(f,read_only=True,data_only=True); ws=wb.active; headers=[norm(str(x.value or '').replace('\u200c','').replace('\u200f','').replace('\u200e','').strip()) for x in next(ws.iter_rows(max_row=1))]
   aliases={'نام':'name','نام و نام خانوادگی':'name','کد دانش‌آموزی':'code','کد':'code','کد ملی':'national_id','پایه':'grade','کلاس':'class_name','شماره تماس':'phone','شماره ولی':'guardian_phone','نام پدر':'father_name','نام پدر و نام خانوادگی پدر':'father_name','شماره پدر':'father_phone','شماره تماس پدر':'father_phone','شماره مادر':'mother_phone','شماره تماس مادر':'mother_phone','واتساپ':'whatsapp','واتساپ پدر':'father_whatsapp','بله پدر':'father_bale','روبیکا پدر':'father_rubika','ایتا پدر':'father_eitaa','واتساپ مادر':'mother_whatsapp','بله مادر':'mother_bale','روبیکا مادر':'mother_rubika','ایتا مادر':'mother_eitaa','بله':'bale','روبیکا':'rubika','ایتا':'eitaa','تاریخ تولد':'birth_date','یادداشت':'notes'}; mapped=[aliases.get(h,h) for h in headers]
   if 'name' not in mapped or 'code' not in mapped: raise ValueError('ستون‌های نام و کد دانش‌آموزی الزامی هستند.')
   c=conn(); added=updated=errors=0; err=[]
   fields=['national_id','grade','class_name','phone','guardian_phone','father_name','father_phone','mother_phone','father_whatsapp','father_bale','father_rubika','father_eitaa','mother_whatsapp','mother_bale','mother_rubika','mother_eitaa','whatsapp','bale','rubika','eitaa','birth_date','notes']
   for i,row in enumerate(ws.iter_rows(min_row=2,values_only=True),2):
    v=dict(zip(mapped,row)); name=excel_text(v.get('name')); code=excel_text(v.get('code'))
    if not name or not code: continue
    birth=excel_text(v.get('birth_date'))
    if birth:
     try: birth=j_to_iso(birth)
     except: errors+=1; err.append(f'ردیف {i}: تاریخ تولد نامعتبر'); continue
    data=[name,code]
    for x in fields:
     if x=='birth_date': data.append(birth)
     elif x in ('phone','guardian_phone','father_phone','mother_phone','whatsapp'): data.append(normalize_phone(v.get(x)))
     else: data.append(excel_text(v.get(x)))
    try:
     old=c.execute('SELECT id FROM students WHERE code=?',(code,)).fetchone()
     if old:
      c.execute('UPDATE students SET name=?,code=?,national_id=?,grade=?,class_name=?,phone=?,guardian_phone=?,father_name=?,father_phone=?,mother_phone=?,father_whatsapp=?,father_bale=?,father_rubika=?,father_eitaa=?,mother_whatsapp=?,mother_bale=?,mother_rubika=?,mother_eitaa=?,whatsapp=?,bale=?,rubika=?,eitaa=?,birth_date=?,notes=? WHERE id=?',tuple(data)+(old['id'],)); updated+=1
     else:
      c.execute('INSERT INTO students(name,code,national_id,grade,class_name,phone,guardian_phone,father_name,father_phone,mother_phone,father_whatsapp,father_bale,father_rubika,father_eitaa,mother_whatsapp,mother_bale,mother_rubika,mother_eitaa,whatsapp,bale,rubika,eitaa,birth_date,notes) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',data); added+=1
    except Exception as e: errors+=1; err.append(f'ردیف {i}: {e}')
   c.commit(); c.close(); wb.close(); result=(added,updated,errors,err); flash(f'ورود انجام شد: {added} جدید، {updated} به‌روزرسانی، {errors} خطا.','success' if not errors else 'warning')
  except Exception as e: flash(f'خطا: {e}','danger')
 return render_template('import_excel.html',result=result)
@app.route('/students/template')
@login_required
def excel_template():
 wb=Workbook(); ws=wb.active; ws.append(['نام','کد دانش‌آموزی','کد ملی','پایه','کلاس','شماره تماس دانش‌آموز','شماره ولی','نام پدر','شماره پدر','شماره مادر','واتساپ پدر','بله پدر','روبیکا پدر','ایتا پدر','واتساپ مادر','بله مادر','روبیکا مادر','ایتا مادر','واتساپ دانش‌آموز','بله دانش‌آموز','روبیکا دانش‌آموز','ایتا دانش‌آموز','تاریخ تولد','یادداشت']); bio=BytesIO(); wb.save(bio); bio.seek(0); return send_file(bio,as_attachment=True,download_name='قالب_دانش_آموزان.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ---------- Bulk messaging ----------
def phone(s): return norm(s).replace(' ','').replace('-','').replace('+','').replace('(','').replace(')','')
def selected_students(ids,grade,class_name):
 c=conn()
 if ids:
  ph=','.join('?'*len(ids)); rows=c.execute(f'SELECT * FROM students WHERE id IN ({ph}) ORDER BY grade,class_name,name',ids).fetchall()
 else:
  rows=c.execute("SELECT * FROM students WHERE (?='' OR grade=?) AND (?='' OR class_name=?) ORDER BY grade,class_name,name",(grade,grade,class_name,class_name)).fetchall()
 c.close(); return rows
@app.route('/messaging',methods=['GET','POST'])
@perm('admin','operator')
def messaging():
    c=conn(); grades=c.execute("SELECT DISTINCT grade FROM students WHERE grade<>'' ORDER BY grade").fetchall(); classes=c.execute("SELECT DISTINCT class_name FROM students WHERE class_name<>'' ORDER BY class_name").fetchall(); all_students=c.execute("SELECT * FROM students ORDER BY grade,class_name,name").fetchall(); c.close()
    links=[]; results=[]; form={}; selected=[]
    if request.method=='POST':
        ids=[int(x) for x in request.form.getlist('student_ids') if str(x).isdigit()]
        grade=request.form.get('grade','').strip(); class_name=request.form.get('class_name','').strip(); target_mode=request.form.get('target_mode','filtered')
        recipient=request.form.get('recipient','both'); platform=request.form.get('platform','whatsapp'); message=request.form.get('message','').strip(); action=request.form.get('action','prepare')
        if target_mode=='selected':
            if not ids: flash('در حالت «فقط دانش‌آموزان انتخاب‌شده»، حداقل یک دانش‌آموز را انتخاب کنید.','warning'); selected=[]
            else: selected=selected_students(ids,'','')
        else: selected=selected_students([],grade,class_name)
        for st in selected:
            recs=[]
            if recipient in ('father','both') and st['father_phone']: recs.append(('پدر',st['father_phone'],'father'))
            if recipient in ('mother','both') and st['mother_phone']: recs.append(('مادر',st['mother_phone'],'mother'))
            if recipient=='student' and st['phone']: recs.append(('دانش‌آموز',st['phone'],'student'))
            text=message.replace('{نام}',st['name']).replace('{پایه}',st['grade'] or '').replace('{کلاس}',st['class_name'] or '').replace('{نام پدر}',st['father_name'] or '')
            for who,num,role in recs:
                if action=='send' and platform=='whatsapp':
                    ok,detail=send_whatsapp(num,text); results.append({'student':st['name'],'who':who,'ok':ok,'detail':detail})
                    c=conn(); c.execute('INSERT INTO message_logs(student_id,recipient,platform,message,status,detail) VALUES(?,?,?,?,?,?)',(st['id'],who,platform,text,'sent' if ok else 'failed',detail)); c.commit(); c.close()
                else:
                    if platform=='whatsapp':
                        contact_field={'father':'father_whatsapp','mother':'mother_whatsapp','student':'whatsapp'}[role]
                        ident=st[contact_field] or num; p=phone(ident); p='98'+p[1:] if p.startswith('0') else p; links.append({'student':st['name'],'who':who,'url':f'https://wa.me/{p}?text={quote(text)}'})
                    else:
                        ident=st[f'{role}_{platform}'] if role in ('father','mother') else st[platform]
                        if ident: links.append({'student':st['name'],'who':who,'url':ident if ident.startswith('http') else {'bale':'https://ble.ir/','rubika':'https://rubika.ir/','eitaa':'https://eitaa.com/'}[platform]+ident.lstrip('@/')})
        if action=='send':
            good=sum(x['ok'] for x in results); flash(f'ارسال خودکار: {good} موفق از {len(results)} مورد.','success' if good==len(results) else 'warning')
        form={'grade':grade,'class_name':class_name,'recipient':recipient,'platform':platform,'message':message,'ids':ids,'target_mode':target_mode}
    return render_template('messaging.html',grades=grades,classes=classes,all_students=all_students,links=links,results=results,form=form,selected=selected,whatsapp_ready=bool(get_messaging_config() and get_messaging_config()['whatsapp_enabled']))


@app.route('/message-history')
@perm('admin','operator','viewer')
def message_history():
 c=conn(); rows=c.execute("SELECT ml.*,COALESCE(s.name,'حذف‌شده') student_name FROM message_logs ml LEFT JOIN students s ON s.id=ml.student_id ORDER BY ml.id DESC LIMIT 1000").fetchall(); c.close(); return render_template('message_history.html',rows=rows)

@app.route('/messaging/settings',methods=['GET','POST'])
@perm('admin')
def messaging_settings():
 if request.method=='POST': save_messaging_config(request.form); flash('تنظیمات واتساپ ذخیره شد.','success'); return redirect(url_for('messaging_settings'))
 return render_template('messaging_settings.html',config=get_messaging_config())

# ---------- Birthdays ----------
@app.route('/birthdays')
@perm('admin','operator','viewer')
def birthdays(): return render_template('birthdays.html',rows=birthday_rows(),reminders=birthday_reminders(),today=today_j())
@app.route('/birthdays/<int:sid>/send', methods=['GET'])
@perm('admin','operator')
def birthday_send_page(sid):
 c=conn(); st=c.execute('SELECT * FROM students WHERE id=?',(sid,)).fetchone(); c.close()
 if not st: abort(404)
 return render_template('birthday_send_form.html',student=st)

@app.post('/birthdays/<int:sid>/send')
@perm('admin','operator')
def send_birthday(sid):
 c=conn(); st=c.execute('SELECT * FROM students WHERE id=?',(sid,)).fetchone(); c.close()
 if not st: abort(404)
 platform=request.form.get('platform','whatsapp').strip().lower(); recipient=request.form.get('recipient','student').strip().lower(); msg=request.form.get('message','').strip()
 if not msg: flash('متن تبریک را وارد کنید.','danger'); return redirect(url_for('birthday_send_page',sid=sid))
 if recipient=='father': num=st['father_phone']; ident=st['father_'+platform] if platform!='whatsapp' else num; label='پدر'
 elif recipient=='mother': num=st['mother_phone']; ident=st['mother_'+platform] if platform!='whatsapp' else num; label='مادر'
 else: num=st['phone']; ident=st[platform] if platform!='whatsapp' else num; label='دانش‌آموز'
 if not ident: flash(f'اطلاعات {label} برای {platform} ثبت نشده است.','danger'); return redirect(url_for('birthday_send_page',sid=sid))
 if platform=='whatsapp':
  if request.form.get('send_now')=='1' and get_messaging_config() and get_messaging_config()['whatsapp_enabled']:
   ok,detail=send_whatsapp(num,msg); c=conn(); c.execute('INSERT INTO message_logs(student_id,recipient,platform,message,status,detail) VALUES(?,?,?,?,?,?)',(st['id'],label,'whatsapp',msg,'sent' if ok else 'failed',detail)); c.commit(); c.close(); flash('تبریک با موفقیت ارسال شد.' if ok else f'ارسال ناموفق: {detail}','success' if ok else 'danger'); return redirect(url_for('birthdays'))
  p=phone(num); p='98'+p[1:] if p.startswith('0') else p; url=f'https://wa.me/{p}?text={quote(msg)}'
 else:
  url=ident if str(ident).startswith('http') else {'bale':'https://ble.ir/','rubika':'https://rubika.ir/','eitaa':'https://eitaa.com/'}[platform]+str(ident).lstrip('@/')
 return render_template('birthday_send.html',student=st,platform=platform,recipient=label,message=msg,url=url,api_ready=bool(get_messaging_config() and get_messaging_config()['whatsapp_enabled'] and platform=='whatsapp'),recipient_key=recipient)

@app.post('/birthdays/<int:sid>/approve')
@perm('admin')
def approve_birthday(sid):
 c=conn(); st=c.execute('SELECT id FROM students WHERE id=?',(sid,)).fetchone()
 if not st: c.close(); abort(404)
 c.execute('UPDATE students SET birthday_approved_year=? WHERE id=?',(today_j()[:4],sid)); c.commit(); c.close(); flash('تولد دانش‌آموز تأیید شد؛ ارسال تبریک همچنان دستی است.','success'); return redirect(url_for('birthdays'))

@app.route('/attendance/report')
@perm('admin','operator','viewer')
def attendance_report():
 start=request.args.get('start',today_j()); end=request.args.get('end',today_j()); grade=request.args.get('grade',''); class_name=request.args.get('class_name',''); q=request.args.get('q','')
 try: si,ei=j_to_iso(start),j_to_iso(end)
 except Exception as e: flash(str(e),'danger'); si=ei=date.today().isoformat()
 c=conn(); students=c.execute("SELECT * FROM students WHERE (?='' OR grade=?) AND (?='' OR class_name=?) AND (name LIKE ? OR code LIKE ? OR ?='') ORDER BY grade,class_name,name",(grade,grade,class_name,class_name,f'%{q}%',f'%{q}%',q)).fetchall();
 rows=[]
 for st in students:
  a=c.execute("SELECT COUNT(*) n FROM attendance WHERE student_id=? AND att_date BETWEEN ? AND ? AND status='غایب'",(st['id'],si,ei)).fetchone()['n']; l=c.execute("SELECT COUNT(*) n FROM attendance WHERE student_id=? AND att_date BETWEEN ? AND ? AND status='تأخیر'",(st['id'],si,ei)).fetchone()['n']; p=c.execute("SELECT COUNT(*) n FROM attendance WHERE student_id=? AND att_date BETWEEN ? AND ? AND status='حاضر'",(st['id'],si,ei)).fetchone()['n']; details=c.execute("SELECT att_date,status FROM attendance WHERE student_id=? AND att_date BETWEEN ? AND ? ORDER BY att_date",(st['id'],si,ei)).fetchall(); rows.append({'student':st,'absent':a,'late':l,'present':p,'details':details})
 grades=c.execute("SELECT DISTINCT grade FROM students WHERE grade<>'' ORDER BY grade").fetchall(); classes=c.execute("SELECT DISTINCT class_name FROM students WHERE class_name<>'' ORDER BY class_name").fetchall(); c.close(); return render_template('attendance_report.html',rows=rows,start=start,end=end,grade=grade,class_name=class_name,q=q,grades=grades,classes=classes)

# ---------- Attendance / Discipline ----------
@app.route('/attendance',methods=['GET','POST'])
@perm('admin','operator','viewer')
def attendance():
 day=request.values.get('day') or today_j(); iso=j_to_iso(day)
 if request.method=='POST' and session['role']!='viewer':
  c=conn()
  for k,v in request.form.items():
   if k.startswith('student_'): c.execute("INSERT INTO attendance(student_id,att_date,status) VALUES(?,?,?) ON CONFLICT(student_id,att_date) DO UPDATE SET status=excluded.status",(int(k[8:]),iso,v))
  c.commit(); c.close(); flash('حضور و غیاب ذخیره شد.','success')
 c=conn(); ss=c.execute('SELECT * FROM students ORDER BY class_name,name').fetchall(); rows=c.execute('SELECT student_id,status FROM attendance WHERE att_date=?',(iso,)).fetchall(); c.close(); return render_template('attendance.html',students=ss,statuses={r['student_id']:r['status'] for r in rows},day=day)
@app.route('/discipline',methods=['GET','POST'])
@perm('admin','operator','viewer')
def discipline():
 if request.method=='POST' and session['role']!='viewer':
  d=request.form; c=conn(); c.execute('INSERT INTO discipline(student_id,event_date,event_type,action,description) VALUES(?,?,?,?,?)',(d['student_id'],j_to_iso(d['event_date']),d['event_type'],d.get('action',''),d.get('description',''))); c.commit(); c.close(); flash('مورد انضباطی ثبت شد.','success'); return redirect(url_for('discipline'))
 c=conn(); rows=c.execute('SELECT d.*,s.name,s.code,s.class_name FROM discipline d JOIN students s ON s.id=d.student_id ORDER BY d.event_date DESC').fetchall(); ss=c.execute('SELECT * FROM students ORDER BY name').fetchall(); c.close(); return render_template('discipline.html',rows=rows,students=ss,today=today_j())
@app.post('/discipline/<int:did>/delete')
@perm('admin')
def delete_discipline(did): c=conn(); c.execute('DELETE FROM discipline WHERE id=?',(did,)); c.commit(); c.close(); flash('حذف شد.','success'); return redirect(url_for('discipline'))


# ---------- Online exams ----------
@app.route('/exams')
@perm('admin','operator','viewer')
def exams():
 c=conn(); rows=c.execute('SELECT e.*,COUNT(q.id) qcount,(SELECT COUNT(*) FROM attempts a WHERE a.exam_id=e.id) acount FROM exams e LEFT JOIN questions q ON q.exam_id=e.id GROUP BY e.id ORDER BY e.id DESC').fetchall(); c.close(); return render_template('exams.html',exams=rows)

@app.route('/exams/new',methods=['GET','POST'])
@perm('admin','operator')
def new_exam():
 if request.method=='POST':
  d=request.form; token=secrets.token_urlsafe(12); c=conn(); cur=c.execute('INSERT INTO exams(title,grade,class_name,start_date,end_date,duration_minutes,phone_required,phone_target,random_questions,created_by,public_token) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(d['title'],d.get('grade',''),d.get('class_name',''),j_to_iso(d['start_date']) if d.get('start_date') else '',j_to_iso(d['end_date']) if d.get('end_date') else '',int(d.get('duration_minutes') or 30),int(d.get('phone_required')=='1'),d.get('phone_target','student'),int(d.get('random_questions')=='1'),session['user_id'],token)); eid=cur.lastrowid
  qs=[]
  for i in range(1,51):
   text=d.get(f'q{i}','').strip()
   if not text: continue
   qs.append((eid,text,d.get(f'type{i}','mcq'),d.get(f'a{i}',''),d.get(f'b{i}',''),d.get(f'c{i}',''),d.get(f'd{i}',''),d.get(f'correct{i}',''),float(d.get(f'points{i}') or 1),i))
  c.executemany('INSERT INTO questions(exam_id,question_text,qtype,option_a,option_b,option_c,option_d,correct_answer,points,sort_order) VALUES(?,?,?,?,?,?,?,?,?,?)',qs); c.commit(); c.close(); flash('آزمون ساخته شد. لینک آماده ارسال است.','success'); return redirect(url_for('exam_detail',eid=eid))
 return render_template('exam_new.html')

@app.route('/exams/<int:eid>')
@perm('admin','operator','viewer')
def exam_detail(eid):
 c=conn(); e=c.execute('SELECT * FROM exams WHERE id=?',(eid,)).fetchone(); qs=c.execute('SELECT * FROM questions WHERE exam_id=? ORDER BY sort_order',(eid,)).fetchall(); attempts=c.execute('SELECT a.*,s.name,s.code FROM attempts a JOIN students s ON s.id=a.student_id WHERE a.exam_id=? ORDER BY a.id DESC',(eid,)).fetchall(); c.close();
 if not e: abort(404)
 return render_template('exam_detail.html',exam=e,questions=qs,attempts=attempts,public_url=url_for('take_exam',token=e['public_token'],_external=True))

@app.route('/exam/<token>',methods=['GET','POST'])
def take_exam(token):
 c=conn(); e=c.execute('SELECT * FROM exams WHERE public_token=?',(token,)).fetchone();
 if not e: c.close(); abort(404)
 now=date.today();
 if e['start_date'] and e['start_date'][:10]>now.isoformat(): c.close(); return render_template('exam_closed.html',message='زمان شروع آزمون نرسیده است.')
 if e['end_date'] and e['end_date'][:10]<now.isoformat(): c.close(); return render_template('exam_closed.html',message='مهلت آزمون به پایان رسیده است.')
 if request.method=='POST' and request.form.get('step')=='login':
  name=request.form.get('name','').strip(); national=norm(request.form.get('national_id','')).strip(); phonev=normalize_phone(request.form.get('phone',''))
  st=c.execute('SELECT * FROM students WHERE national_id=? AND name=?',(national,name)).fetchone()
  if not st: c.close(); return render_template('exam_login.html',exam=e,error='نام و کد ملی در فهرست این مدرسه پیدا نشد.')
  if e['phone_required'] and not phonev: c.close(); return render_template('exam_login.html',exam=e,error='شماره تلفن برای این آزمون اجباری است.')
  old=c.execute('SELECT id FROM attempts WHERE exam_id=? AND student_id=?',(e['id'],st['id'])).fetchone()
  if old: c.close(); return render_template('exam_closed.html',message='شما قبلاً در این آزمون شرکت کرده‌اید و امکان شرکت مجدد ندارید.')
  aid=c.execute('INSERT INTO attempts(exam_id,student_id,phone,started_at,status) VALUES(?,?,?,?,?)',(e['id'],st['id'],phonev,datetime.now().isoformat(timespec='seconds'),'started')).lastrowid; c.commit(); qs=c.execute('SELECT * FROM questions WHERE exam_id=? ORDER BY sort_order',(e['id'],)).fetchall(); c.close(); return render_template('take_exam.html',exam=e,questions=qs,attempt_id=aid,student=st)
 if request.method=='POST' and request.form.get('step')=='submit':
  aid=int(request.form['attempt_id']); at=c.execute('SELECT * FROM attempts WHERE id=? AND exam_id=?',(aid,e['id'])).fetchone()
  if not at: c.close(); abort(403)
  qs=c.execute('SELECT * FROM questions WHERE exam_id=? ORDER BY sort_order',(e['id'],)).fetchall(); total=0; maxscore=sum(float(q['points']) for q in qs)
  for q in qs:
   ans=request.form.get(f'q_{q["id"]}',''); pts=float(q['points']) if q['qtype']=='mcq' and ans==q['correct_answer'] else 0; total+=pts; c.execute('INSERT INTO answers(attempt_id,question_id,answer,points) VALUES(?,?,?,?)',(aid,q['id'],ans,pts))
  score=round((total/maxscore*10),2) if maxscore else 0; c.execute("UPDATE attempts SET score=?,submitted_at=?,status='submitted' WHERE id=?",(score,datetime.now().isoformat(timespec='seconds'),aid)); c.commit(); st=c.execute('SELECT * FROM students WHERE id=?',(at['student_id'],)).fetchone(); c.close(); return render_template('exam_result.html',exam=e,student=st,score=score,total=total,maxscore=maxscore)
 c.close(); return render_template('exam_login.html',exam=e,error=None)

@app.route('/exams/<int:eid>/send-results',methods=['POST'])
@perm('admin','operator')
def send_exam_results(eid):
 platform=request.form.get('platform','whatsapp'); c=conn(); rows=c.execute('SELECT a.*,s.name,s.father_phone,s.mother_phone,s.phone FROM attempts a JOIN students s ON s.id=a.student_id WHERE a.exam_id=? AND a.status=\'submitted\'',(eid,)).fetchall(); e=c.execute('SELECT title FROM exams WHERE id=?',(eid,)).fetchone(); c.close(); good=bad=0
 for r in rows:
  for who,num in [('پدر',r['father_phone']),('مادر',r['mother_phone'])]:
   if not num: continue
   msg=f"نتیجه آزمون «{e['title']}»: {r['score']} از ۱۰"
   ok,_=send_or_prepare(r,who,platform,msg) if platform=='whatsapp' else (False,'API فعال نیست')
   good+=ok; bad+=not ok
 flash(f'نتایج: {good} ارسال موفق، {bad} ناموفق.','success' if bad==0 else 'warning'); return redirect(url_for('exam_detail',eid=eid))

# ---------- Users / Backup / Reports ----------
@app.route('/users',methods=['GET','POST'])
@perm('admin')
def users():
 if request.method=='POST':
  try:
   c=conn(); c.execute('INSERT INTO users(username,password_hash,role,must_change_password) VALUES(?,?,?,1)',(request.form['username'].strip(),generate_password_hash(request.form['password']),request.form['role'])); c.commit(); c.close(); flash('کاربر ایجاد شد.','success')
  except sqlite3.IntegrityError: flash('نام کاربری تکراری است.','danger')
 c=conn(); rows=c.execute('SELECT * FROM users ORDER BY username').fetchall(); c.close(); return render_template('users.html',users=rows)
@app.post('/users/<int:uid>/toggle')
@perm('admin')
def toggle_user(uid): c=conn(); c.execute('UPDATE users SET active=1-active WHERE id=?',(uid,)); c.commit(); c.close(); return redirect(url_for('users'))
@app.post('/users/<int:uid>/reset')
@perm('admin')
def reset_user(uid): c=conn(); c.execute('UPDATE users SET password_hash=?,must_change_password=1 WHERE id=?',(generate_password_hash('123456'),uid)); c.commit(); c.close(); flash('رمز به 123456 تغییر کرد.','warning'); return redirect(url_for('users'))
@app.route('/backup')
@perm('admin')
def backup(): fn=BACKUPS/f'backup_{today_j().replace("/","-")}_{datetime.now():%H%M%S}.db'; shutil.copy2(DB,fn); return send_file(fn,as_attachment=True,download_name=fn.name)
@app.route('/restore',methods=['GET','POST'])
@perm('admin')
def restore():
 if request.method=='POST':
  f=request.files.get('backup')
  if not f: flash('فایل انتخاب نشده.','danger')
  else:
   with tempfile.NamedTemporaryFile(delete=False,suffix='.db') as t: f.save(t.name); tmp=t.name
   try:
    test=sqlite3.connect(tmp); ok=test.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='students'").fetchone(); test.close()
    if not ok: raise ValueError('فایل پشتیبان معتبر نیست.')
    shutil.copy2(DB,BACKUPS/f'before_restore_{datetime.now():%Y%m%d_%H%M%S}.db'); shutil.copy2(tmp,DB); flash('بازیابی انجام شد.','success')
   except Exception as e: flash(str(e),'danger')
   finally: os.unlink(tmp)
 return render_template('backup.html')
@app.route('/reports')
@perm('admin','operator','viewer')
def reports(): return render_template('reports.html')
@app.route('/reports/students/print')
@perm('admin','operator','viewer')
def print_students():
 c=conn(); rows=c.execute('SELECT * FROM students ORDER BY grade,class_name,name').fetchall(); c.close(); return render_template('print_students.html',students=rows)
@app.route('/reports/attendance/print')
@perm('admin','operator','viewer')
def print_attendance():
 day=request.args.get('day') or today_j(); iso=j_to_iso(day); c=conn(); rows=c.execute("SELECT s.name,s.code,s.class_name,COALESCE(a.status,'ثبت نشده') status FROM students s LEFT JOIN attendance a ON a.student_id=s.id AND a.att_date=? ORDER BY s.class_name,s.name",(iso,)).fetchall(); c.close(); return render_template('print_attendance.html',students=rows,day=day)
@app.route('/reports/discipline/print')
@perm('admin','operator','viewer')
def print_discipline(): c=conn(); rows=c.execute('SELECT d.*,s.name,s.class_name FROM discipline d JOIN students s ON s.id=d.student_id ORDER BY d.event_date DESC').fetchall(); c.close(); return render_template('print_discipline.html',rows=rows)
@app.route('/reports/students.csv')
@perm('admin','operator','viewer')
def students_csv():
 c=conn(); rows=c.execute('SELECT name,code,grade,class_name,phone,guardian_phone,father_phone,mother_phone,birth_date FROM students ORDER BY grade,class_name,name').fetchall(); c.close(); out=StringIO(); w=csv.writer(out); w.writerow(['نام','کد','پایه','کلاس','تماس','ولی','پدر','مادر','تاریخ تولد']); [w.writerow([r['name'],r['code'],r['grade'],r['class_name'],r['phone'],r['guardian_phone'],r['father_name'],r['father_phone'],r['mother_phone'],iso_to_j(r['birth_date'])]) for r in rows]; return send_file(BytesIO(out.getvalue().encode('utf-8-sig')),as_attachment=True,download_name='گزارش_دانش_آموزان.csv',mimetype='text/csv')

# ---------- Single messaging / parent invitation ----------
@app.route('/message/<int:sid>/<platform>')
@login_required
def message(sid,platform):
    role=request.args.get('role','student')
    c=conn(); s=c.execute('SELECT * FROM students WHERE id=?',(sid,)).fetchone(); c.close()
    if not s: abort(404)
    field_map={
      'father':{'whatsapp':'father_whatsapp','bale':'father_bale','rubika':'father_rubika','eitaa':'father_eitaa','phone':'father_phone'},
      'mother':{'whatsapp':'mother_whatsapp','bale':'mother_bale','rubika':'mother_rubika','eitaa':'mother_eitaa','phone':'mother_phone'},
      'student':{'whatsapp':'whatsapp','bale':'bale','rubika':'rubika','eitaa':'eitaa','phone':'phone'}
    }
    role='father' if role=='father' else 'mother' if role=='mother' else 'student'
    field=field_map[role].get(platform)
    v=(s[field] or '') if field else ''
    if not v:
        flash(f'اطلاعات {"پدر" if role=="father" else "مادر" if role=="mother" else "دانش‌آموز"} برای {platform} ثبت نشده است.','warning'); return redirect(url_for('students'))
    if platform=='whatsapp':
        p=phone(v); p='98'+p[1:] if p.startswith('0') else p
        return redirect('https://wa.me/'+p)
    if v.startswith('http'): return redirect(v)
    return redirect({'bale':'https://ble.ir/','rubika':'https://rubika.ir/','eitaa':'https://eitaa.com/'}[platform]+v.lstrip('@/'))

@app.route('/invitation-file/<token>')
def invitation_file(token):
    c=conn(); row=c.execute('SELECT * FROM invitation_files WHERE token=?',(token,)).fetchone(); c.close()
    if not row: abort(404)
    path=INV_UPLOADS/row['stored_name']
    if not path.exists(): abort(404)
    return send_file(path,download_name=row['original_name'],mimetype=row['mime_type'] or None,as_attachment=False)

@app.route('/parent-invitation',methods=['GET','POST'])
@perm('admin','operator')
def parent_invitation():
    c=conn(); students_all=c.execute('SELECT * FROM students ORDER BY grade,class_name,name').fetchall(); c.close()
    selected=[]; invitation={}; links=[]; results=[]; attachment=None
    if request.method=='POST':
        ids=[int(x) for x in request.form.getlist('student_ids') if str(x).isdigit()]
        selected=selected_students(ids,'','')
        invitation={k:request.form.get(k,'').strip() for k in ('title','date','time','place','text','platform','recipient')}; action=request.form.get('action','prepare')
        invitation['title']=invitation['title'] or 'دعوت‌نامه ولی دانش‌آموز'
        invitation['text']=invitation['text'] or 'از شما دعوت می‌شود در برنامه مدرسه حضور داشته باشید.'
        up=request.files.get('invitation_file')
        if up and up.filename:
            ext=Path(up.filename).suffix.lower().lstrip('.')
            if ext not in ALLOWED_INV_FILES:
                flash('نوع فایل مجاز نیست. فقط PNG، JPG، WEBP، PDF، DOC و DOCX مجاز است.','danger')
            else:
                up.stream.seek(0,2); size=up.stream.tell(); up.stream.seek(0)
                if size>MAX_INV_FILE_SIZE:
                    flash('حجم فایل دعوت‌نامه نباید بیشتر از ۱۰ مگابایت باشد.','danger')
                else:
                    token=secrets.token_urlsafe(18); safe=secure_filename(up.filename) or ('invitation.'+ext); stored=f'{token}_{safe}'
                    up.save(INV_UPLOADS/stored)
                    c=conn(); c.execute('INSERT INTO invitation_files(token,original_name,stored_name,mime_type) VALUES(?,?,?,?)',(token,up.filename,stored,up.mimetype)); c.commit(); c.close()
                    attachment={'token':token,'name':up.filename,'url':url_for('invitation_file',token=token,_external=True)}
        old_token=request.form.get('attachment_token','').strip()
        if not attachment and old_token:
            c=conn(); row=c.execute('SELECT * FROM invitation_files WHERE token=?',(old_token,)).fetchone(); c.close()
            if row: attachment={'token':row['token'],'name':row['original_name'],'url':url_for('invitation_file',token=row['token'],_external=True)}
        invitation['attachment_token']=attachment['token'] if attachment else ''
        for st in selected:
            body=(f"{invitation['title']}\n\nاولیای گرامی دانش‌آموز {st['name']}،\n{invitation['text']}\n\nتاریخ: {invitation['date']}\nساعت: {invitation['time']}\nمکان: {invitation['place']}\n")
            if attachment: body += f"\n📎 فایل/پوستر دعوت‌نامه: {attachment['url']}\n"
            body += "\nبا احترام\nمدیریت مدرسه"
            rec=invitation.get('recipient','both'); recs=[]
            if rec in ('father','both') and st['father_phone']: recs.append(('پدر',st['father_phone'],'father'))
            if rec in ('mother','both') and st['mother_phone']: recs.append(('مادر',st['mother_phone'],'mother'))
            for who,num,role in recs:
                platform=invitation.get('platform','whatsapp')
                if action=='send' and platform=='whatsapp':
                    img_path=None
                    if attachment:
                        c2=conn(); ar=c2.execute('SELECT stored_name,mime_type FROM invitation_files WHERE token=?',(attachment['token'],)).fetchone(); c2.close()
                        if ar and (ar['mime_type'] or '').startswith('image/'):
                            img_path=INV_UPLOADS/ar['stored_name']
                    ok,detail=send_whatsapp_image(num,img_path,body) if img_path else send_whatsapp(num,body)
                    results.append({'student':st['name'],'who':who,'ok':ok,'detail':detail})
                    c=conn(); c.execute('INSERT INTO message_logs(student_id,recipient,platform,message,status,detail) VALUES(?,?,?,?,?,?)',(st['id'],who,platform,body,'sent' if ok else 'failed',detail)); c.commit(); c.close()
                if platform=='whatsapp':
                    p=phone(num); p='98'+p[1:] if p.startswith('0') else p; links.append({'student':st['name'],'who':who,'url':f'https://wa.me/{p}?text={quote(body)}','message':body})
                else:
                    key=f'{role}_{platform}'; ident=st[key] or ''
                    if ident:
                        url=ident if ident.startswith('http') else {'bale':'https://ble.ir/','rubika':'https://rubika.ir/','eitaa':'https://eitaa.com/'}[platform]+ident.lstrip('@/')
                        links.append({'student':st['name'],'who':who,'url':url,'message':body})
        if action=='send' and invitation.get('platform')=='whatsapp':
            good=sum(x['ok'] for x in results); flash(f'ارسال دعوت‌نامه: {good} موفق از {len(results)} مورد.','success' if good==len(results) else 'warning')
    return render_template('parent_invitation.html',students=students_all,selected=selected,invitation=invitation,links=links,results=results,attachment=attachment,whatsapp_ready=bool(get_messaging_config() and get_messaging_config()['whatsapp_enabled']))

if __name__=='__main__': app.run(host='127.0.0.1',port=5000,debug=False)
